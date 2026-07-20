from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.template.loader import render_to_string


def export_to_excel(headers, rows, filename='report'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'تقرير'
    ws.sheet_view.rightToLeft = True

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_to_pdf(template_name, context, filename='report'):
    try:
        from weasyprint import HTML
        html_string = render_to_string(template_name, context)
        pdf = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
    except Exception:
        response = HttpResponse('PDF export unavailable', status=500)
        return response
